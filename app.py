
import os, io, csv, re
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, session, flash, Response
from werkzeug.security import check_password_hash
import psycopg
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = os.environ["SECRET_KEY"]
DATABASE_URL = os.environ["DATABASE_URL"]
ADMIN_PASSWORD_HASH = os.environ["ADMIN_PASSWORD_HASH"]

ARABIC_DIGITS = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")

def db():
    return psycopg.connect(DATABASE_URL)

def normalize_digits(value):
    if value is None:
        return ""
    s = str(value).strip().translate(ARABIC_DIGITS)
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return re.sub(r"\D", "", s)

def phone_last4(value):
    digits = normalize_digits(value)
    return digits[-4:] if len(digits) >= 4 else digits

def clean_text(value):
    if value is None:
        return ""
    s = str(value).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s

def init_db():
    with db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS students (
                    student_id TEXT PRIMARY KEY,
                    student_name TEXT NOT NULL,
                    major TEXT NOT NULL,
                    phone_last4 TEXT NOT NULL,
                    lineup_number TEXT NOT NULL,
                    rehearsal_day TEXT,
                    rehearsal_time TEXT,
                    notes TEXT
                )
            """)
            # Upgrade an existing database created by the previous version.
            cur.execute("""
                ALTER TABLE students
                ADD COLUMN IF NOT EXISTS rehearsal_day TEXT
            """)
        conn.commit()

def admin_required(fn):
    @wraps(fn)
    def inner(*args, **kwargs):
        if not session.get("admin"):
            return redirect(url_for("admin_login"))
        return fn(*args, **kwargs)
    return inner

@app.before_request
def ensure_db():
    init_db()

@app.route("/", methods=["GET", "POST"])
def index():
    result = None
    error = None
    if request.method == "POST":
        student_id = clean_text(request.form.get("student_id",""))
        last4 = normalize_digits(request.form.get("phone_last4",""))[-4:]
        if not student_id or len(last4) != 4:
            error = "يرجى إدخال الرقم الجامعي وآخر 4 أرقام من رقم الهاتف."
        else:
            with db() as conn:
                with conn.cursor() as cur:
                    cur.execute("""
                        SELECT student_name, major, lineup_number,
                               COALESCE(rehearsal_day,''),
                               COALESCE(rehearsal_time,''),
                               COALESCE(notes,'')
                        FROM students
                        WHERE student_id=%s AND phone_last4=%s
                    """, (student_id, last4))
                    row = cur.fetchone()
            if row:
                result = {
                    "student_name": row[0],
                    "major": row[1],
                    "lineup_number": row[2],
                    "rehearsal_day": row[3],
                    "rehearsal_time": row[4],
                    "notes": row[5],
                }
            else:
                error = "لم يتم العثور على بيانات مطابقة. يرجى التأكد من الرقم الجامعي وآخر 4 أرقام من الهاتف."
    return render_template("index.html", result=result, error=error)

@app.route("/health")
def health():
    return {"status":"ok"}, 200

@app.route("/admin", methods=["GET","POST"])
def admin_login():
    if session.get("admin"):
        return redirect(url_for("admin_dashboard"))
    if request.method == "POST":
        password = request.form.get("password","")
        if check_password_hash(ADMIN_PASSWORD_HASH, password):
            session["admin"] = True
            return redirect(url_for("admin_dashboard"))
        flash("كلمة المرور غير صحيحة.", "error")
    return render_template("admin_login.html")

@app.route("/admin/dashboard")
@admin_required
def admin_dashboard():
    with db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM students")
            count = cur.fetchone()[0]
            cur.execute("""
                SELECT student_id, student_name, major, lineup_number, rehearsal_day, rehearsal_time
                FROM students ORDER BY student_id LIMIT 100
            """)
            rows = cur.fetchall()
    return render_template("admin_dashboard.html", count=count, rows=rows)

HEADER_ALIASES = {
    "student_id": ["student_id", "student id", "الرقم الجامعي", "رقم الطالب"],
    "student_name": ["student_name", "student name", "اسم الطالب", "الاسم"],
    "major": ["major", "التخصص"],
    "phone_number": ["phone_number", "phone number", "phone", "mobile", "رقم الهاتف", "الهاتف", "رقم الموبايل"],
    "lineup_number": ["lineup_number", "lineup number", "رقم الاصطفاف"],
    "rehearsal_day": ["rehearsal_day", "rehearsal day", "يوم البروفة", "تاريخ البروفة"],
    "rehearsal_time": ["rehearsal_time", "rehearsal time", "وقت البروفة", "موعد البروفة"],
    "notes": ["notes", "ملاحظات", "الملاحظات"],
}

def canonical_headers(headers):
    lookup = {}
    normalized = {clean_text(h).lower(): h for h in headers if h is not None}
    for canonical, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            key = alias.lower()
            if key in normalized:
                lookup[canonical] = normalized[key]
                break
    return lookup

def read_xlsx(file_storage):
    data = io.BytesIO(file_storage.read())
    wb = load_workbook(data, read_only=True, data_only=True)
    ws = wb.active
    values = list(ws.iter_rows(values_only=True))
    if not values:
        return [], {}
    headers = [clean_text(x) for x in values[0]]
    mapping = canonical_headers(headers)
    records = []
    for row in values[1:]:
        item = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        records.append(item)
    return records, mapping

def read_csv(file_storage):
    content = file_storage.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    headers = reader.fieldnames or []
    mapping = canonical_headers(headers)
    return list(reader), mapping

@app.route("/admin/upload", methods=["POST"])
@admin_required
def admin_upload():
    f = request.files.get("file")
    if not f or not f.filename:
        flash("يرجى اختيار ملف Excel أو CSV.", "error")
        return redirect(url_for("admin_dashboard"))

    filename = f.filename.lower()
    try:
        if filename.endswith(".xlsx"):
            records, mapping = read_xlsx(f)
        elif filename.endswith(".csv"):
            records, mapping = read_csv(f)
        else:
            flash("الملف يجب أن يكون بصيغة .xlsx أو .csv", "error")
            return redirect(url_for("admin_dashboard"))

        required = ["student_id","student_name","major","phone_number","lineup_number"]
        missing = [x for x in required if x not in mapping]
        if missing:
            flash("الملف لا يحتوي جميع الأعمدة المطلوبة: " + ", ".join(missing), "error")
            return redirect(url_for("admin_dashboard"))

        prepared = []
        for r in records:
            sid = clean_text(r.get(mapping["student_id"]))
            if not sid:
                continue
            last4 = phone_last4(r.get(mapping["phone_number"]))
            if len(last4) != 4:
                raise ValueError(f"رقم الهاتف غير صالح للطالب {sid}")

            prepared.append((
                sid,
                clean_text(r.get(mapping["student_name"])),
                clean_text(r.get(mapping["major"])),
                last4,
                clean_text(r.get(mapping["lineup_number"])),
                clean_text(r.get(mapping.get("rehearsal_day"))) if mapping.get("rehearsal_day") else "",
                clean_text(r.get(mapping.get("rehearsal_time"))) if mapping.get("rehearsal_time") else "",
                clean_text(r.get(mapping.get("notes"))) if mapping.get("notes") else "",
            ))

        with db() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM students")
                cur.executemany("""
                    INSERT INTO students
                    (student_id, student_name, major, phone_last4, lineup_number,
                     rehearsal_day, rehearsal_time, notes)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                """, prepared)
            conn.commit()

        flash(f"تم تحميل {len(prepared)} طالبًا بنجاح. تم استخراج آخر 4 أرقام من الهاتف تلقائيًا.", "success")

    except Exception as e:
        flash(f"تعذر قراءة الملف: {e}", "error")

    return redirect(url_for("admin_dashboard"))

@app.route("/admin/sample")
@admin_required
def sample():
    csv_text = """student_id,student_name,major,phone_number,lineup_number,rehearsal_day,rehearsal_time,notes
202012345,طالب تجريبي,المحاسبة,0791236789,145,الثلاثاء 25/08/2026,10:00 صباحاً,يرجى الحضور قبل الموعد بـ 30 دقيقة
"""
    return Response(csv_text, mimetype="text/csv; charset=utf-8",
                    headers={"Content-Disposition":"attachment; filename=students_template.csv"})

@app.route("/admin/logout")
@admin_required
def logout():
    session.clear()
    return redirect(url_for("admin_login"))
